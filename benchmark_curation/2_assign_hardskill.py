"""
Hard Skill 分配主程序
为已生成的员工数据分配专业技能（Hard Skills）
"""

import os
import sys
import json
import time
from typing import List, Dict, Tuple, Optional
from datetime import datetime

# 第三方库
try:
    from openai import OpenAI
    from openpyxl import load_workbook, Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from tqdm import tqdm
except ImportError as e:
    print(f"错误: 缺少必要的依赖库 - {e}")
    print("请运行: pip install openai openpyxl python-dotenv tqdm")
    sys.exit(1)

# 导入配置和提示词
import config
from prompt import PromptTemplate


# ========== 数据加载模块 ==========

def read_employee_data(file_path: str) -> List[Dict]:
    """
    从Excel读取员工数据

    参数:
        file_path: Excel文件路径

    返回:
        员工数据列表
    """
    print(f"\n{'='*60}")
    print("加载员工数据...")
    print(f"{'='*60}")

    if not os.path.exists(file_path):
        print(f"✗ 文件不存在: {file_path}")
        return None

    try:
        wb = load_workbook(file_path)
        ws = wb.active

        employees = []

        # 读取表头
        headers = {}
        for col_idx, cell in enumerate(ws[1], start=1):
            headers[cell.value] = col_idx

        # 验证必要的列
        required_cols = ["Name", "User_ID", "Team", "Dept_ID", "Rank", "Title"]
        for col in required_cols:
            if col not in headers:
                print(f"✗ 缺少必要的列: {col}")
                return None

        # 读取数据（从第2行开始）
        for row_idx in tqdm(range(2, ws.max_row + 1), desc="读取员工数据", unit="行"):
            employee = {
                "name": ws.cell(row=row_idx, column=headers["Name"]).value,
                "user_id": ws.cell(row=row_idx, column=headers["User_ID"]).value,
                "team": ws.cell(row=row_idx, column=headers["Team"]).value,
                "dept_id": ws.cell(row=row_idx, column=headers["Dept_ID"]).value,
                "rank": ws.cell(row=row_idx, column=headers["Rank"]).value,
                "title": ws.cell(row=row_idx, column=headers["Title"]).value,
            }
            employees.append(employee)

        print(f"✓ 成功加载 {len(employees)} 名员工数据")
        return employees

    except Exception as e:
        print(f"✗ 读取文件失败: {e}")
        return None


# ========== 技能集合生成模块 ==========

def extract_unique_titles(employees: List[Dict]) -> List[str]:
    """
    提取所有唯一职位

    参数:
        employees: 员工数据列表

    返回:
        唯一职位列表
    """
    titles = set(emp.get("title", "") for emp in employees)
    titles = sorted([t for t in titles if t])  # 移除空值并排序
    print(f"\n提取到 {len(titles)} 个唯一职位")
    return titles


def call_gpt5(prompt: str, retries: int = None) -> Optional[Dict]:
    """
    调用 OpenAI GPT-5 API

    参数:
        prompt: 提示词
        retries: 重试次数（默认使用配置值）

    返回:
        解析后的 JSON 数据，失败返回 None
    """
    if retries is None:
        retries = config.MAX_RETRIES

    # 检查 API Key
    if not config.OPENAI_API_KEY:
        print("错误: OPENAI_API_KEY 未设置")
        return None

    # 初始化 OpenAI 客户端（通过 OpenRouter）
    client = OpenAI(
        api_key=config.OPENAI_API_KEY,
        base_url=config.OPENAI_BASE_URL
    )

    for attempt in range(retries):
        try:
            print(f"\n{'='*60}")
            print(f"正在调用 OpenAI API (尝试 {attempt + 1}/{retries})...")
            print(f"{'='*60}")

            # 准备 API 参数
            api_kwargs = {
                "model": config.OPENAI_MODEL,
                "messages": [
                    {
                        "role": "system",
                        "content": "你是一位专业的人力资源和技能管理专家。请严格按照要求输出 JSON 格式的数据。"
                    },
                    {
                        "role": "user",
                        "content": prompt
                    }
                ],
            }

            # 添加配置中的参数
            api_kwargs.update(config.API_PARAMS)

            # 显示等待信息
            print("⏳ 正在等待 GPT-5 生成数据...")
            print("   请耐心等待，不要中断程序...\n")

            response = client.chat.completions.create(**api_kwargs)

            # 提取响应内容
            content = response.choices[0].message.content

            # 检查是否为空
            if content is None:
                print(f"✗ API 返回 None 内容")
                print(f"使用的 tokens: {response.usage.total_tokens}")
                if attempt < retries - 1:
                    print(f"将在 {config.RETRY_DELAY} 秒后重试...")
                    time.sleep(config.RETRY_DELAY)
                    continue
                else:
                    return None

            content = content.strip()

            if config.VERBOSE:
                print(f"\nAPI 响应长度: {len(content)} 字符")
                print(f"使用的 tokens: {response.usage.total_tokens}")
                if len(content) == 0:
                    print(f"⚠ 警告：API 返回空字符串")
                    print(f"finish_reason: {response.choices[0].finish_reason}")

            # 如果内容为空，重试
            if len(content) == 0:
                print(f"✗ API 返回空内容")
                if attempt < retries - 1:
                    print(f"将在 {config.RETRY_DELAY} 秒后重试...")
                    time.sleep(config.RETRY_DELAY)
                    continue
                else:
                    return None

            # 清理响应内容（移除可能的 markdown 标记）
            if content.startswith("```json"):
                content = content[7:]
            if content.startswith("```"):
                content = content[3:]
            if content.endswith("```"):
                content = content[:-3]
            content = content.strip()

            # 解析 JSON
            try:
                data = json.loads(content)
                print("✓ API 调用成功，数据解析完成")
                return data
            except json.JSONDecodeError as e:
                print(f"✗ JSON 解析失败: {e}")
                print(f"响应内容（前500字符）: {content[:500]}")

                if attempt < retries - 1:
                    print(f"将在 {config.RETRY_DELAY} 秒后重试...")
                    time.sleep(config.RETRY_DELAY)
                    continue
                else:
                    return None

        except Exception as e:
            print(f"✗ API 调用失败: {e}")

            if attempt < retries - 1:
                print(f"将在 {config.RETRY_DELAY} 秒后重试...")
                time.sleep(config.RETRY_DELAY)
                continue
            else:
                return None

    return None


def generate_skill_universe(unique_titles: List[str]) -> Optional[Dict]:
    """
    生成闭合的技能集合

    参数:
        unique_titles: 唯一职位列表

    返回:
        技能数据字典，包含 skill_universe 和 validation 信息
    """
    print(f"\n{'='*60}")
    print("生成技能集合...")
    print(f"{'='*60}")

    # 检查是否已存在技能集合文件
    if os.path.exists(config.SKILL_UNIVERSE_FILE):
        print(f"发现已存在的技能集合文件: {config.SKILL_UNIVERSE_FILE}")
        try:
            with open(config.SKILL_UNIVERSE_FILE, 'r', encoding='utf-8') as f:
                response_data = json.load(f)

            # 扁平化技能集合
            flat_skills = []
            for category_data in response_data.get("skill_universe", []):
                if "skills" in category_data:
                    flat_skills.extend(category_data["skills"])

            total_count = len(flat_skills)
            print(f"✓ 已加载现有技能集合，共 {total_count} 个技能")

            return {
                "skill_universe": flat_skills,
                "total_count": total_count,
                "validation": response_data.get("validation", {}),
                "raw_data": response_data
            }
        except Exception as e:
            print(f"⚠ 读取技能集合文件失败: {e}")
            print("将通过 API 重新生成...")

    # 生成提示词
    prompt = PromptTemplate.get_skill_universe_prompt(unique_titles, config.TEAMS)

    if config.VERBOSE:
        print(f"提示词长度: {len(prompt)} 字符")

    # 调用 GPT-5
    response_data = call_gpt5(prompt)

    if not response_data:
        print("✗ 技能集合生成失败")
        return None

    # 验证响应格式
    if "skill_universe" not in response_data:
        print("✗ 响应数据格式错误：缺少 'skill_universe' 字段")
        return None

    # 扁平化技能集合（从分类结构转为简单列表）
    flat_skills = []
    for category_data in response_data["skill_universe"]:
        if "skills" in category_data:
            flat_skills.extend(category_data["skills"])

    total_count = len(flat_skills)
    print(f"✓ 成功生成技能集合，共 {total_count} 个技能")

    # 保存到文件（可选）
    if config.SKILL_UNIVERSE_FILE:
        try:
            with open(config.SKILL_UNIVERSE_FILE, 'w', encoding='utf-8') as f:
                json.dump(response_data, f, ensure_ascii=False, indent=2)
            print(f"✓ 技能集合已保存到: {config.SKILL_UNIVERSE_FILE}")
        except Exception as e:
            print(f"⚠ 保存技能集合文件失败: {e}")

    return {
        "skill_universe": flat_skills,
        "total_count": total_count,
        "validation": response_data.get("validation", {}),
        "raw_data": response_data
    }


def validate_mutual_exclusivity(skills: List[str]) -> Tuple[bool, List[str]]:
    """
    验证技能互斥性（本地验证）

    参数:
        skills: 技能列表

    返回:
        (是否通过, 错误信息列表)
    """
    errors = []

    # 规则1: 检查重复（大小写不敏感）
    normalized = [s.lower().strip() for s in skills]
    if len(normalized) != len(set(normalized)):
        # 找出重复项
        seen = set()
        duplicates = set()
        for s in normalized:
            if s in seen:
                duplicates.add(s)
            seen.add(s)
        errors.append(f"存在重复技能: {duplicates}")

    # 规则2: 检查明显的父子关系（简单示例）
    parent_child_patterns = [
        ("python", "python3"),
        ("java", "java8"),
        ("javascript", "js"),
        ("excel", "microsoft excel"),
    ]

    for parent, child in parent_child_patterns:
        has_parent = any(parent == s.lower() for s in skills)
        has_child = any(child == s.lower() for s in skills)
        if has_parent and has_child:
            errors.append(f"存在父子关系: {parent} 和 {child}")

    return len(errors) == 0, errors


# ========== 技能分配模块 ==========

def assign_skills_in_batches(employees: List[Dict],
                              skill_universe: List[str],
                              batch_size: int = 20) -> Optional[Dict[str, List[Dict]]]:
    """
    批量分配技能

    参数:
        employees: 员工数据列表
        skill_universe: 可用技能集合
        batch_size: 每批分配的员工数量

    返回:
        技能分配字典 {员工姓名: [{"skill": ..., "level": ...}, ...]}
    """
    print(f"\n{'='*60}")
    print(f"开始分批分配技能（每批{batch_size}个）...")
    print(f"{'='*60}")

    assignments = {}
    total_batches = (len(employees) + batch_size - 1) // batch_size

    with tqdm(total=len(employees), desc="总体进度", unit="员工") as pbar:
        for batch_num in range(1, total_batches + 1):
            # 计算当前批次的员工
            start_idx = (batch_num - 1) * batch_size
            end_idx = min(start_idx + batch_size, len(employees))
            batch_employees = employees[start_idx:end_idx]

            print(f"\n--- 批次 {batch_num}/{total_batches} ---")
            print(f"分配 {len(batch_employees)} 名员工的技能...")

            # 生成提示词
            prompt = PromptTemplate.get_skill_assignment_prompt(
                batch_employees,
                skill_universe,
                batch_num
            )

            if config.VERBOSE:
                print(f"提示词长度: {len(prompt)} 字符")

            # 调用 GPT API
            response_data = call_gpt5(prompt)

            if not response_data:
                print(f"✗ 批次 {batch_num} 技能分配失败")
                return None

            # 提取分配数据
            if "assignments" not in response_data:
                print(f"✗ 批次 {batch_num} 响应格式错误：缺少 'assignments' 字段")
                return None

            batch_assignments = response_data["assignments"]

            # 存储到结果字典
            for assignment in batch_assignments:
                name = assignment.get("name", "")
                hard_skills = assignment.get("hard_skills", [])
                assignments[name] = hard_skills

            print(f"✓ 批次 {batch_num} 完成，已分配 {len(batch_assignments)} 名员工的技能")
            print(f"  累计: {len(assignments)}/{len(employees)}")

            # 更新进度条
            pbar.update(len(batch_assignments))

    print(f"\n✓ 所有批次完成，共为 {len(assignments)} 名员工分配技能")
    return assignments


def merge_skills_to_employees(employees: List[Dict],
                               assignments: Dict[str, List[Dict]]) -> List[Dict]:
    """
    将技能分配合并到员工数据

    参数:
        employees: 员工数据列表
        assignments: 技能分配字典

    返回:
        合并后的员工数据列表
    """
    print(f"\n{'='*60}")
    print("合并技能数据...")
    print(f"{'='*60}")

    for employee in tqdm(employees, desc="合并数据", unit="员工"):
        name = employee.get("name", "")
        if name in assignments:
            employee["hard_skills"] = assignments[name]
        else:
            employee["hard_skills"] = []
            print(f"⚠ 员工 {name} 未找到技能分配")

    print(f"✓ 数据合并完成")
    return employees


# ========== 数据验证模块 ==========

def validate_skill_assignments(employees: List[Dict],
                                 skill_universe: List[str]) -> Tuple[bool, List[str]]:
    """
    验证技能分配的正确性

    参数:
        employees: 员工数据列表（包含技能）
        skill_universe: 可用技能集合

    返回:
        (是否验证通过, 错误信息列表)
    """
    print(f"\n{'='*60}")
    print("验证技能分配...")
    print(f"{'='*60}")

    errors = []
    skill_universe_set = set(skill_universe)

    for employee in tqdm(employees, desc="验证数据", unit="员工"):
        name = employee.get("name", "")
        rank = employee.get("rank", 0)
        hard_skills = employee.get("hard_skills", [])

        # 1. 检查是否有技能
        if config.SKILL_VALIDATION_RULES["check_all_employees_have_skills"]:
            if not hard_skills:
                errors.append(f"{name}: 没有技能分配")
                continue

        # 2. 检查技能数量
        if config.SKILL_VALIDATION_RULES["check_skill_count_range"]:
            skill_count = len(hard_skills)
            min_count = config.SKILL_COUNT_BY_RANK.get(rank, {}).get("min", 0)
            max_count = config.SKILL_COUNT_BY_RANK.get(rank, {}).get("max", 100)

            if not (min_count <= skill_count <= max_count):
                errors.append(
                    f"{name}: 技能数量 {skill_count} 不在范围内 ({min_count}-{max_count})"
                )

        # 3. 检查技能是否在集合内
        if config.SKILL_VALIDATION_RULES["check_skills_in_universe"]:
            for skill_item in hard_skills:
                skill = skill_item.get("skill", "")
                if skill not in skill_universe_set:
                    errors.append(f"{name}: 技能 '{skill}' 不在技能集合中")

        # 4. 检查员工内部技能重复
        if config.SKILL_VALIDATION_RULES["check_no_duplicates_per_employee"]:
            skills_only = [s.get("skill", "") for s in hard_skills]
            if len(skills_only) != len(set(skills_only)):
                errors.append(f"{name}: 存在重复技能")

        # 5. 检查技能强度等级有效性
        if config.SKILL_VALIDATION_RULES["check_skill_levels_valid"]:
            for skill_item in hard_skills:
                level = skill_item.get("level", "")
                if level not in config.SKILL_LEVELS:
                    errors.append(
                        f"{name}: 技能强度等级 '{level}' 无效（应为 strong/medium/low）"
                    )

        # 6. 检查强度分布（至少有1个strong）
        if config.SKILL_VALIDATION_RULES["check_level_distribution"]:
            levels = [s.get("level", "") for s in hard_skills]
            if "strong" not in levels:
                errors.append(f"{name}: 没有 strong 等级的技能")

    # 输出验证结果
    if errors:
        print(f"✗ 数据验证失败，发现 {len(errors)} 个问题:")
        for error in errors[:10]:  # 只显示前10个错误
            print(f"  - {error}")
        if len(errors) > 10:
            print(f"  ... 还有 {len(errors) - 10} 个错误")
        return False, errors
    else:
        print("✓ 数据验证通过")
        return True, []


# ========== Excel 输出模块 ==========

def write_to_excel_with_skills(employees: List[Dict], output_path: str) -> bool:
    """
    写入带技能的Excel文件

    参数:
        employees: 员工数据列表（包含技能）
        output_path: 输出文件路径

    返回:
        是否成功
    """
    print(f"\n{'='*60}")
    print("写入 Excel 文件...")
    print(f"{'='*60}")

    try:
        # 创建工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = "员工技能信息"

        # 定义样式
        header_font = Font(bold=True, size=12, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        cell_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

        # 写入表头
        headers = {
            "name": "Name",
            "user_id": "User_ID",
            "team": "Team",
            "dept_id": "Dept_ID",
            "rank": "Rank",
            "title": "Title",
            "hard_skills": "Hard_Skills"
        }

        for col_idx, (key, header) in enumerate(headers.items(), start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # 写入数据
        for row_idx, employee in tqdm(enumerate(employees, start=2), total=len(employees), desc="写入Excel数据", unit="行"):
            for col_idx, key in enumerate(headers.keys(), start=1):
                if key == "hard_skills":
                    # 格式化技能列: "Python(strong), Docker(medium), MySQL(low)"
                    skills = employee.get(key, [])
                    if skills:
                        skill_strs = [f"{s['skill']}({s['level']})" for s in skills]
                        value = ", ".join(skill_strs)
                    else:
                        value = ""
                else:
                    value = employee.get(key, "")

                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = cell_alignment

        # 调整列宽
        column_widths = {
            1: 15,  # Name
            2: 12,  # User_ID
            3: 18,  # Team
            4: 12,  # Dept_ID
            5: 8,   # Rank
            6: 25,  # Title
            7: 60   # Hard_Skills (更宽以容纳技能列表)
        }

        for col_idx, width in column_widths.items():
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = width

        # 保存文件
        print("\n正在保存 Excel 文件...")
        with tqdm(total=100, desc="保存文件", bar_format='{l_bar}{bar}| {n_fmt}/{total_fmt}') as pbar:
            wb.save(output_path)
            pbar.update(100)

        print(f"✓ Excel 文件已保存: {output_path}")
        return True

    except Exception as e:
        print(f"✗ 写入 Excel 文件失败: {e}")
        return False


# ========== 统计分析模块 ==========

def analyze_skill_distribution(employees: List[Dict]) -> Dict:
    """
    分析技能分布

    参数:
        employees: 员工数据列表

    返回:
        统计信息字典
    """
    from collections import Counter

    # 技能使用频率
    skill_counter = Counter()
    level_counter = Counter()
    skills_by_rank = {1: [], 2: [], 3: []}

    for employee in employees:
        rank = employee.get("rank", 0)
        hard_skills = employee.get("hard_skills", [])

        for skill_item in hard_skills:
            skill = skill_item.get("skill", "")
            level = skill_item.get("level", "")

            skill_counter[skill] += 1
            level_counter[level] += 1

            if rank in skills_by_rank:
                skills_by_rank[rank].append(skill)

    return {
        "skill_frequency": skill_counter,
        "level_distribution": level_counter,
        "skills_by_rank": skills_by_rank
    }


def print_skill_statistics(employees: List[Dict], skill_universe: List[str]):
    """
    打印技能统计信息

    参数:
        employees: 员工数据列表
        skill_universe: 技能集合
    """
    print(f"\n{'='*60}")
    print("技能统计信息")
    print(f"{'='*60}")

    stats = analyze_skill_distribution(employees)

    # 技能集合大小
    print(f"\n技能集合总数: {len(skill_universe)}")

    # 技能使用频率 Top 10
    print(f"\n技能使用频率 Top 10:")
    for skill, count in stats["skill_frequency"].most_common(10):
        print(f"  {skill}: {count} 人")

    # 强度等级分布
    print(f"\n技能强度等级分布:")
    for level in ["strong", "medium", "low"]:
        count = stats["level_distribution"].get(level, 0)
        print(f"  {level}: {count} 次")

    # 每个职别的平均技能数
    print(f"\n每个职别的平均技能数:")
    for rank in [1, 2, 3]:
        employees_with_rank = [e for e in employees if e.get("rank") == rank]
        if employees_with_rank:
            avg_skills = sum(len(e.get("hard_skills", [])) for e in employees_with_rank) / len(employees_with_rank)
            rank_name = config.RANKS.get(rank, f"Rank {rank}")
            print(f"  {rank_name}: {avg_skills:.1f} 个")

    # 示例员工（前5个）
    print(f"\n示例员工（前5个）:")
    for i, employee in enumerate(employees[:5], start=1):
        print(f"\n  [{i}] {employee.get('name', 'N/A')} - {employee.get('title', 'N/A')}")
        hard_skills = employee.get("hard_skills", [])
        if hard_skills:
            for skill_item in hard_skills:
                print(f"      • {skill_item['skill']} ({skill_item['level']})")
        else:
            print("      （无技能）")

    print(f"{'='*60}")


# ========== 主程序 ==========

def main():
    """主程序入口"""
    print("\n" + "="*60)
    print("Hard Skill 分配系统")
    print("="*60)
    print(f"启动时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")

    # 1. 加载员工数据
    employees = read_employee_data(config.EMPLOYEE_FILE_PATH)
    if not employees:
        print("\n✗ 加载员工数据失败")
        return 1

    # 2. 生成技能集合
    unique_titles = extract_unique_titles(employees)
    skill_data = generate_skill_universe(unique_titles)
    if not skill_data:
        print("\n✗ 生成技能集合失败")
        return 1

    skill_universe = skill_data["skill_universe"]

    # 验证互斥性（本地验证）
    is_valid, errors = validate_mutual_exclusivity(skill_universe)
    if not is_valid:
        print("\n⚠ 技能集合互斥性验证警告:")
        for error in errors:
            print(f"  - {error}")

    # 3. 批量分配技能
    assignments = assign_skills_in_batches(
        employees,
        skill_universe,
        batch_size=config.SKILL_ASSIGNMENT_BATCH_SIZE
    )
    if not assignments:
        print("\n✗ 技能分配失败")
        return 1

    # 4. 合并数据
    employees = merge_skills_to_employees(employees, assignments)

    # 5. 验证数据
    is_valid, errors = validate_skill_assignments(employees, skill_universe)
    if not is_valid:
        print(f"\n⚠ 发现 {len(errors)} 个验证错误")
        # 继续执行，但输出警告

    # 6. 输出Excel
    success = write_to_excel_with_skills(employees, config.OUTPUT_PATH_WITH_HARDSKILLS)
    if not success:
        return 1

    # 7. 统计报告
    print_skill_statistics(employees, skill_universe)

    # 完成
    print(f"\n{'='*60}")
    print("✓ 所有任务完成!")
    print(f"{'='*60}")
    print(f"输出文件: {config.OUTPUT_PATH_WITH_HARDSKILLS}")
    print(f"完成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"{'='*60}\n")

    return 0


if __name__ == "__main__":
    exit_code = main()
    sys.exit(exit_code)
