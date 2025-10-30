"""
员工数据生成主程序
通过调用 OpenAI GPT-5 API 生成员工信息并输出为 Excel 文件
"""

import os
import sys
import json
import random
import string
import time
from typing import List, Dict, Tuple, Optional
from datetime import datetime

# 第三方库
try:
    from openai import OpenAI
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from tqdm import tqdm
except ImportError as e:
    print(f"错误: 缺少必要的依赖库 - {e}")
    print("请运行: pip install openai openpyxl python-dotenv tqdm")
    sys.exit(1)

# 导入配置和提示词
import config
from prompt import PromptTemplate


# ========== ID 生成模块 ==========

def generate_user_id(existing_ids: set = None) -> str:
    """
    生成唯一的 User_ID

    参数:
        existing_ids: 已存在的ID集合，确保不重复

    返回:
        8位随机字符串
    """
    if existing_ids is None:
        existing_ids = set()

    while True:
        user_id = ''.join(random.choices(config.USER_ID_CHARS, k=config.USER_ID_LENGTH))
        if user_id not in existing_ids:
            existing_ids.add(user_id)
            return user_id


def generate_dept_id(existing_ids: set = None) -> str:
    """
    生成唯一的 Dept_ID

    参数:
        existing_ids: 已存在的ID集合，确保不重复

    返回:
        5位随机字符串
    """
    if existing_ids is None:
        existing_ids = set()

    while True:
        dept_id = ''.join(random.choices(config.DEPT_ID_CHARS, k=config.DEPT_ID_LENGTH))
        if dept_id not in existing_ids:
            existing_ids.add(dept_id)
            return dept_id


def generate_dept_id_mapping(teams: List[str]) -> Dict[str, str]:
    """
    为所有团队生成 Dept_ID 映射

    参数:
        teams: 团队列表

    返回:
        团队名到Dept_ID的映射字典
    """
    dept_ids = set()
    mapping = {}

    for team in teams:
        mapping[team] = generate_dept_id(dept_ids)

    # 为大领导的"高层管理"也生成一个ID
    mapping["高层管理"] = generate_dept_id(dept_ids)

    return mapping


# ========== OpenAI API 调用模块 ==========

def call_gpt5(prompt: str, retries: int = None) -> Optional[Dict]:
    """
    调用 OpenAI GPT-5 API

    参数:
        prompt: 提示词
        retries: 重试次数（默认使用配置值）

    返回:
        解析后的 JSON 数据，失败返回 None
    """
    if retries is None:
        retries = config.MAX_RETRIES

    # 检查 API Key
    if not config.OPENAI_API_KEY:
        print("错误: OPENAI_API_KEY 未设置")
        print("请在 .env 文件中设置: OPENAI_API_KEY=your_api_key")
        return None

    # 初始化 OpenAI 客户端（通过 OpenRouter）
    client = OpenAI(
        api_key=config.OPENAI_API_KEY,
        base_url=config.OPENAI_BASE_URL
    )

    for attempt in range(retries):
        try:
            print(f"\n{'='*60}")
            print(f"正在调用 OpenAI API (尝试 {attempt + 1}/{retries})...")
            print(f"{'='*60}")

            # 调用 API
            # 准备 API 参数（只传递配置中存在的参数）
            api_kwargs = {
                "model": config.OPENAI_MODEL,
                "messages": [
                    {
                        "role": "system",
                        "content": "你是一位专业的人力资源专家，擅长设计组织架构和生成员工信息。请严格按照要求输出 JSON 格式的数据。"
                    },
                    {
                        "role": "user",
                        "content": prompt
                    }
                ],
            }

            # 添加配置中的参数
            api_kwargs.update(config.API_PARAMS)

            # 显示等待信息（API 调用是阻塞操作，可能需要1-3分钟）
            print("⏳ 正在等待 GPT-5 生成数据...")
            print(f"   预计需要 1-3 分钟（生成 {config.TOTAL_EMPLOYEES} 个员工）")
            print("   请耐心等待，不要中断程序...\n")

            response = client.chat.completions.create(**api_kwargs)

            # 提取响应内容
            content = response.choices[0].message.content.strip()

            if config.VERBOSE:
                print(f"\nAPI 响应长度: {len(content)} 字符")
                print(f"使用的 tokens: {response.usage.total_tokens}")

            # 清理响应内容（移除可能的 markdown 标记）
            if content.startswith("```json"):
                content = content[7:]
            if content.startswith("```"):
                content = content[3:]
            if content.endswith("```"):
                content = content[:-3]
            content = content.strip()

            # 解析 JSON
            try:
                data = json.loads(content)
                print("✓ API 调用成功，数据解析完成")
                return data
            except json.JSONDecodeError as e:
                print(f"✗ JSON 解析失败: {e}")
                print(f"响应内容（前500字符）: {content[:500]}")

                if attempt < retries - 1:
                    print(f"将在 {config.RETRY_DELAY} 秒后重试...")
                    time.sleep(config.RETRY_DELAY)
                    continue
                else:
                    return None

        except Exception as e:
            print(f"✗ API 调用失败: {e}")

            if attempt < retries - 1:
                print(f"将在 {config.RETRY_DELAY} 秒后重试...")
                time.sleep(config.RETRY_DELAY)
                continue
            else:
                return None

    return None


# ========== 数据生成模块 ==========

def generate_employee_data() -> Optional[List[Dict]]:
    """
    生成完整的员工数据

    返回:
        员工数据列表，失败返回 None
    """
    print(f"\n{'='*60}")
    print("开始生成员工数据...")
    print(f"{'='*60}")

    # 生成提示词
    prompt = PromptTemplate.get_main_prompt(
        total_count=config.TOTAL_EMPLOYEES,
        teams=config.TEAMS
    )

    if config.VERBOSE:
        print(f"\n提示词长度: {len(prompt)} 字符")

    # 调用 GPT-5
    response_data = call_gpt5(prompt)

    if not response_data:
        print("✗ 数据生成失败")
        return None

    # 提取员工数据
    if "employees" not in response_data:
        print("✗ 响应数据格式错误：缺少 'employees' 字段")
        return None

    employees = response_data["employees"]

    print(f"✓ 成功生成 {len(employees)} 名员工数据")

    return employees


def generate_employee_data_in_batches(batch_size: int = 20) -> Optional[List[Dict]]:
    """
    分批生成员工数据

    参数:
        batch_size: 每批生成的员工数量（默认20）

    返回:
        所有员工数据列表，失败返回 None
    """
    print(f"\n{'='*60}")
    print(f"开始分批生成员工数据（每批{batch_size}个）...")
    print(f"{'='*60}")

    all_employees = []
    existing_names = []
    total_batches = (config.TOTAL_EMPLOYEES + batch_size - 1) // batch_size

    # 计算第一批的数量（包含领导）
    first_batch_size = config.TOP_LEADER_COUNT + config.TEAM_LEADER_COUNT + (batch_size - 8)
    if first_batch_size > config.TOTAL_EMPLOYEES:
        first_batch_size = config.TOTAL_EMPLOYEES

    # 使用tqdm显示总体进度
    with tqdm(total=config.TOTAL_EMPLOYEES, desc="总体进度", unit="员工") as pbar:
        for batch_num in range(1, total_batches + 1):
            # 确定当前批次大小
            if batch_num == 1:
                current_batch_size = first_batch_size
                needs_leaders = True
            else:
                remaining = config.TOTAL_EMPLOYEES - len(all_employees)
                current_batch_size = min(batch_size, remaining)
                needs_leaders = False

            if current_batch_size <= 0:
                break

            print(f"\n--- 批次 {batch_num}/{total_batches} ---")
            print(f"生成 {current_batch_size} 名员工...")

            # 生成提示词
            prompt = PromptTemplate.get_batch_prompt(
                batch_num=batch_num,
                batch_size=current_batch_size,
                teams=config.TEAMS,
                existing_names=existing_names,
                needs_leaders=needs_leaders
            )

            if config.VERBOSE:
                print(f"提示词长度: {len(prompt)} 字符")

            # 调用 GPT API
            response_data = call_gpt5(prompt)

            if not response_data:
                print(f"✗ 批次 {batch_num} 生成失败")
                return None

            # 提取员工数据
            if "employees" not in response_data:
                print(f"✗ 批次 {batch_num} 响应格式错误：缺少 'employees' 字段")
                return None

            batch_employees = response_data["employees"]

            if len(batch_employees) != current_batch_size:
                print(f"⚠ 批次 {batch_num} 生成数量不符：期望{current_batch_size}，实际{len(batch_employees)}")

            # 添加到总列表
            all_employees.extend(batch_employees)
            existing_names.extend([e.get("name", "") for e in batch_employees])

            print(f"✓ 批次 {batch_num} 完成，已生成 {len(batch_employees)} 名员工")
            print(f"  累计: {len(all_employees)}/{config.TOTAL_EMPLOYEES}")

            # 更新进度条
            pbar.update(len(batch_employees))

    print(f"\n✓ 所有批次完成，共生成 {len(all_employees)} 名员工")
    return all_employees


def add_ids_to_employees(employees: List[Dict], dept_mapping: Dict[str, str]) -> List[Dict]:
    """
    为员工数据添加 User_ID 和 Dept_ID

    参数:
        employees: 员工数据列表
        dept_mapping: 部门ID映射

    返回:
        添加了ID的员工数据列表
    """
    print(f"\n{'='*60}")
    print("生成 User_ID 和 Dept_ID...")
    print(f"{'='*60}")

    user_ids = set()

    # 使用 tqdm 显示进度
    for employee in tqdm(employees, desc="生成员工ID", unit="员工"):
        # 生成 User_ID
        employee["user_id"] = generate_user_id(user_ids)

        # 添加 Dept_ID
        team = employee.get("team", "")
        employee["dept_id"] = dept_mapping.get(team, "UNKNOWN")

    print(f"✓ 已为 {len(employees)} 名员工生成ID")

    return employees


# ========== 数据修复模块 ==========

def generate_chinese_name(existing_names: set = None) -> str:
    """
    生成随机的中文姓名（本地生成，避免重复）

    参数:
        existing_names: 已存在的姓名集合

    返回:
        唯一的中文姓名
    """
    if existing_names is None:
        existing_names = set()

    # 常见姓氏（100个）
    surnames = [
        "王", "李", "张", "刘", "陈", "杨", "黄", "赵", "吴", "周",
        "徐", "孙", "马", "朱", "胡", "郭", "何", "高", "林", "罗",
        "郑", "梁", "谢", "宋", "唐", "许", "韩", "冯", "邓", "曹",
        "彭", "曾", "肖", "田", "董", "袁", "潘", "于", "蒋", "蔡",
        "余", "杜", "叶", "程", "苏", "魏", "吕", "丁", "任", "沈",
        "姚", "卢", "姜", "崔", "钟", "谭", "陆", "汪", "范", "金",
        "石", "廖", "贾", "夏", "韦", "付", "方", "白", "邹", "孟",
        "熊", "秦", "邱", "江", "尹", "薛", "闫", "段", "雷", "侯",
        "龙", "史", "陶", "黎", "贺", "顾", "毛", "郝", "龚", "邵",
        "万", "钱", "严", "覃", "武", "戴", "莫", "孔", "向", "汤"
    ]

    # 常见名字用字（单字名和双字名）
    name_chars = [
        "伟", "芳", "娜", "秀", "敏", "静", "丽", "强", "磊", "军",
        "洋", "勇", "艳", "杰", "涛", "明", "超", "秀英", "娟", "英",
        "华", "慧", "巧", "美", "娜", "倩", "妮", "莉", "峰", "辉",
        "刚", "平", "辉", "鹏", "宇", "晨", "帆", "航", "凯", "斌",
        "文", "博", "浩", "昊", "然", "睿", "轩", "宸", "泽", "瑞",
        "琪", "瑶", "萱", "怡", "涵", "晴", "雪", "梦", "婷", "欣",
        "雨", "馨", "悦", "诗", "嘉", "颖", "洁", "雅", "薇", "莹",
        "佳", "思", "宏", "建", "国", "志", "东", "波", "鑫", "亮",
        "玲", "红", "霞", "春", "兰", "燕", "云", "新", "琳", "敏",
        "婕", "晗", "曦", "羽", "昕", "妍", "彤", "萌", "蕾", "璐"
    ]

    # 尝试生成唯一姓名（最多1000次）
    for _ in range(1000):
        surname = random.choice(surnames)
        # 70%概率生成双字名，30%概率生成单字名
        if random.random() < 0.7:
            given_name = random.choice(name_chars) + random.choice(name_chars)
        else:
            given_name = random.choice(name_chars)

        full_name = surname + given_name

        if full_name not in existing_names:
            return full_name

    # 如果还是重复，使用数字后缀
    for i in range(10000):
        surname = random.choice(surnames)
        given_name = random.choice(name_chars)
        full_name = f"{surname}{given_name}{i}" if i > 0 else f"{surname}{given_name}"
        if full_name not in existing_names:
            return full_name

    raise Exception("无法生成唯一姓名，请检查数据")


def fix_duplicate_names(employees: List[Dict]) -> Tuple[List[Dict], int]:
    """
    自动修复重复的姓名

    参数:
        employees: 员工数据列表

    返回:
        (修复后的员工列表, 修复的数量)
    """
    print(f"\n{'='*60}")
    print("开始修复重复姓名...")
    print(f"{'='*60}")

    # 统计姓名出现次数
    name_counts = {}
    for employee in employees:
        name = employee.get("name", "")
        name_counts[name] = name_counts.get(name, 0) + 1

    # 找出重复的姓名
    duplicate_names = {name for name, count in name_counts.items() if count > 1}

    if not duplicate_names:
        print("✓ 没有发现重复姓名")
        return employees, 0

    print(f"发现 {len(duplicate_names)} 个重复姓名: {duplicate_names}")

    # 收集所有已存在的姓名
    existing_names = set(employee.get("name", "") for employee in employees)

    fixed_count = 0

    # 对每个重复姓名，保留第一个出现的，修改后续出现的
    for duplicate_name in duplicate_names:
        found_first = False

        for i, employee in enumerate(employees):
            if employee.get("name") == duplicate_name:
                if not found_first:
                    # 保留第一个
                    found_first = True
                    print(f"  保留第一个: {duplicate_name}")
                else:
                    # 生成新姓名
                    new_name = generate_chinese_name(existing_names)
                    existing_names.add(new_name)

                    print(f"  修复: {duplicate_name} → {new_name}")
                    print(f"       (职位: {employee.get('title', 'N/A')}, 团队: {employee.get('team', 'N/A')})")

                    employees[i]["name"] = new_name
                    fixed_count += 1

    print(f"\n✓ 修复完成，共修复 {fixed_count} 个重复姓名")

    return employees, fixed_count


# ========== 数据验证模块 ==========

def validate_employee_data(employees: List[Dict]) -> Tuple[bool, List[str]]:
    """
    验证员工数据的完整性和正确性

    参数:
        employees: 员工数据列表

    返回:
        (是否验证通过, 错误信息列表)
    """
    print(f"\n{'='*60}")
    print("验证数据...")
    print(f"{'='*60}")

    errors = []

    # 1. 检查总数
    if config.VALIDATION_RULES["check_total_count"]:
        if len(employees) != config.TOTAL_EMPLOYEES:
            errors.append(f"员工总数错误: 期望 {config.TOTAL_EMPLOYEES}，实际 {len(employees)}")

    # 2. 检查姓名唯一性
    if config.VALIDATION_RULES["check_unique_names"]:
        names = [e.get("name", "") for e in employees]
        unique_names = set(names)
        if len(unique_names) != len(names):
            duplicates = [name for name in names if names.count(name) > 1]
            errors.append(f"姓名重复: {set(duplicates)}")

    # 3. 检查 User_ID 唯一性
    if config.VALIDATION_RULES["check_unique_user_ids"]:
        user_ids = [e.get("user_id", "") for e in employees]
        unique_user_ids = set(user_ids)
        if len(unique_user_ids) != len(user_ids):
            errors.append("User_ID 存在重复")

    # 4. 检查职别分布
    if config.VALIDATION_RULES["check_rank_distribution"]:
        rank_counts = {1: 0, 2: 0, 3: 0}
        for e in employees:
            rank = e.get("rank", 0)
            if rank in rank_counts:
                rank_counts[rank] += 1

        for rank, expected_count in config.RANK_COUNTS.items():
            actual_count = rank_counts[rank]
            if actual_count != expected_count:
                errors.append(
                    f"Rank {rank} 数量错误: 期望 {expected_count}，实际 {actual_count}"
                )

    # 5. 检查团队覆盖
    if config.VALIDATION_RULES["check_team_coverage"]:
        teams_in_data = set(e.get("team", "") for e in employees)
        for team in config.TEAMS:
            if team not in teams_in_data:
                errors.append(f"团队 '{team}' 没有员工")

    # 6. 检查必填字段
    required_fields = ["name", "user_id", "team", "dept_id", "rank", "title"]
    for i, employee in tqdm(enumerate(employees), total=len(employees), desc="验证数据完整性", unit="记录"):
        for field in required_fields:
            if field not in employee or not employee[field]:
                errors.append(f"第 {i+1} 条记录缺少字段: {field}")
                break

    # 输出验证结果
    if errors:
        print("✗ 数据验证失败:")
        for error in errors:
            print(f"  - {error}")
        return False, errors
    else:
        print("✓ 数据验证通过")
        return True, []


# ========== Excel 写入模块 ==========

def write_to_excel(employees: List[Dict], output_path: str) -> bool:
    """
    将员工数据写入 Excel 文件

    参数:
        employees: 员工数据列表
        output_path: 输出文件路径

    返回:
        是否成功
    """
    print(f"\n{'='*60}")
    print("写入 Excel 文件...")
    print(f"{'='*60}")

    try:
        # 创建工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = "员工信息"

        # 定义样式
        header_font = Font(bold=True, size=12, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        cell_alignment = Alignment(horizontal="left", vertical="center")

        # 写入表头
        headers = {
            "name": "Name",
            "user_id": "User_ID",
            "team": "Team",
            "dept_id": "Dept_ID",
            "rank": "Rank",
            "title": "Title"
        }

        for col_idx, (key, header) in enumerate(headers.items(), start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # 写入数据
        for row_idx, employee in tqdm(enumerate(employees, start=2), total=len(employees), desc="写入Excel数据", unit="行"):
            for col_idx, key in enumerate(headers.keys(), start=1):
                value = employee.get(key, "")
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = cell_alignment

        # 调整列宽
        column_widths = {
            1: 15,  # Name
            2: 12,  # User_ID
            3: 18,  # Team
            4: 12,  # Dept_ID
            5: 8,   # Rank
            6: 25   # Title
        }

        for col_idx, width in column_widths.items():
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = width

        # 保存文件
        print("\n正在保存 Excel 文件...")
        with tqdm(total=100, desc="保存文件", bar_format='{l_bar}{bar}| {n_fmt}/{total_fmt}') as pbar:
            wb.save(output_path)
            pbar.update(100)

        print(f"✓ Excel 文件已保存: {output_path}")
        return True

    except Exception as e:
        print(f"✗ 写入 Excel 文件失败: {e}")
        return False


# ========== 辅助函数 ==========

def create_output_directory():
    """创建输出目录"""
    if config.CREATE_OUTPUT_DIR:
        os.makedirs(config.OUTPUT_DIR, exist_ok=True)


def print_summary(employees: List[Dict], dept_mapping: Dict[str, str]):
    """打印数据摘要"""
    print(f"\n{'='*60}")
    print("数据摘要")
    print(f"{'='*60}")

    # 统计职别分布
    rank_dist = {1: 0, 2: 0, 3: 0}
    for e in employees:
        rank = e.get("rank", 0)
        if rank in rank_dist:
            rank_dist[rank] += 1

    print(f"总员工数: {len(employees)}")
    print(f"  - Rank 1 (大领导): {rank_dist[1]}")
    print(f"  - Rank 2 (团队领导): {rank_dist[2]}")
    print(f"  - Rank 3 (普通员工): {rank_dist[3]}")

    # 统计团队分布
    print(f"\n团队分布:")
    team_dist = {}
    for e in employees:
        team = e.get("team", "")
        team_dist[team] = team_dist.get(team, 0) + 1

    for team in sorted(team_dist.keys()):
        count = team_dist[team]
        dept_id = dept_mapping.get(team, "N/A")
        print(f"  - {team} (ID: {dept_id}): {count} 人")

    # 显示示例数据
    print(f"\n示例数据（前3条）:")
    for i, employee in enumerate(employees[:3], start=1):
        print(f"\n  [{i}] {employee.get('name', 'N/A')}")
        print(f"      User_ID: {employee.get('user_id', 'N/A')}")
        print(f"      Team: {employee.get('team', 'N/A')}")
        print(f"      Dept_ID: {employee.get('dept_id', 'N/A')}")
        print(f"      Rank: {employee.get('rank', 'N/A')}")
        print(f"      Title: {employee.get('title', 'N/A')}")

    print(f"{'='*60}")


# ========== 主程序 ==========

def main():
    """主程序入口"""
    print("\n" + "="*60)
    print("员工数据生成系统")
    print("="*60)
    print(f"启动时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")

    # 打印配置摘要
    config.print_config_summary()

    # 验证配置
    config_errors = config.validate_config()
    if config_errors:
        print("\n配置错误:")
        for error in config_errors:
            print(f"  ✗ {error}")
        return 1

    # 创建输出目录
    create_output_directory()

    # 生成部门ID映射
    print(f"\n{'='*60}")
    print("生成部门ID映射...")
    print(f"{'='*60}")
    dept_mapping = generate_dept_id_mapping(config.TEAMS)
    for team, dept_id in dept_mapping.items():
        print(f"  {team}: {dept_id}")

    # 生成员工数据（使用分批生成）
    employees = generate_employee_data_in_batches(batch_size=20)
    if not employees:
        print("\n✗ 员工数据生成失败")
        return 1

    # 添加 ID
    employees = add_ids_to_employees(employees, dept_mapping)

    # 验证数据
    is_valid, errors = validate_employee_data(employees)

    if not is_valid:
        # 检查是否只是姓名重复的问题
        has_name_duplicates = any("姓名重复" in error for error in errors)
        other_errors = [error for error in errors if "姓名重复" not in error]

        if has_name_duplicates and not other_errors:
            # 只有姓名重复问题，尝试自动修复
            print("\n⚠ 检测到姓名重复，正在自动修复...")
            employees, fixed_count = fix_duplicate_names(employees)

            # 重新验证
            print("\n重新验证数据...")
            is_valid, errors = validate_employee_data(employees)

            if not is_valid:
                print("\n✗ 修复后仍有错误，程序终止")
                return 1
            else:
                print(f"\n✓ 自动修复成功！已修复 {fixed_count} 个重复姓名")
        else:
            # 有其他错误
            if config.STRICT_VALIDATION:
                print("\n✗ 数据验证失败，程序终止")
                return 1
            else:
                print("\n⚠ 数据验证失败，但继续执行（严格验证已禁用）")

    # 打印数据摘要
    print_summary(employees, dept_mapping)

    # 写入 Excel
    success = write_to_excel(employees, config.OUTPUT_PATH)

    if not success:
        return 1

    # 完成
    print(f"\n{'='*60}")
    print("✓ 所有任务完成!")
    print(f"{'='*60}")
    print(f"输出文件: {config.OUTPUT_PATH}")
    print(f"完成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"{'='*60}\n")

    return 0


if __name__ == "__main__":
    exit_code = main()
    sys.exit(exit_code)
